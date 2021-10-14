from rest_framework.exceptions import ParseError
from rest_framework.response import Response
from .serializers import FileSerializer
from rest_framework import status, parsers, renderers
from rest_framework.generics import GenericAPIView
from openpyxl import load_workbook


def get_summary(ws, columns_found, summary):
    append_values = False
    sum = 0
    counter = 0
    for col_idx in range(ws.max_column):
        for row in range(1, ws.max_row+1):
            if type(ws[row][col_idx].value) == str and ws[row][col_idx].value.strip() in columns_found.keys() and append_values is False:
                col_name = ws[row][col_idx].value.strip()
                columns_found[col_name] = True
                result_dict = {"column": col_name}
                sum = 0
                counter = 0
                append_values = True
            if append_values is True:
                if type(ws[row][col_idx].value) in [float, int]:
                    sum += ws[row][col_idx].value
                    counter += 1
                if (row == ws.max_row or type(ws[row+1][col_idx].value) is str) and counter > 0:
                    result_dict["sum"] = round(sum, 2)
                    result_dict["avg"] = round(sum/counter, 2)
                    append_values = False
                    summary.append(result_dict)

        append_values = False


class FileUploadView(GenericAPIView):
    parser_classes = (
        parsers.FormParser,
        parsers.MultiPartParser,
        parsers.FileUploadParser,
    )
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = FileSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            data = serializer.validated_data
            file = data['file']
            wb = load_workbook(file)
            columns = set(data['columns'][0].split(','))
            columns_found = {column: False for column in columns}
            summary = []
            for sheet in wb:
                get_summary(sheet, columns_found, summary)
            for key, val in columns_found.items():
                if val and key not in [key["column"] for key in summary]:
                    summary.append({"column": key, "info": "column doesn't have numeric values"})
                if not val:
                    summary.append({"column": key, "info": "column not found"})
            return Response({
                'file': str(file),
                'summary': summary
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)